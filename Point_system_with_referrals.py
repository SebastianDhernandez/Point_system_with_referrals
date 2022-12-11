# Necessary libraries
from openpyxl import load_workbook  # Library for work with excel sheets

# Filepath of the excel file
filepath = r"Path" # Example C:\Users\user\Downloads\sample.xlsx



# Class for create agents
class Agent:
    def __init__(self, name):
        self.Name = name
        self.__Sons1 = []
        self.__Sons2 = []
        self.__Sons3 = []

    # Function to change the sons
    def changesons1(self, son):
        if son != self.Name:
            if son not in self.__Sons2:
                if son not in self.__Sons3:
                    self.__Sons1.append(son)

    def changesons2(self, son):
        if son != self.Name:
            if son not in self.__Sons1:
                if son not in self.__Sons3:
                    self.__Sons2.append(son)

    def changesons3(self, son):
        if son != self.Name:
            if son not in self.__Sons2:
                if son not in self.__Sons1:
                    self.__Sons3.append(son)

    # Function to get the sons
    def getsons1(self):
        return self.__Sons1

    def getsons2(self):
        return self.__Sons2

    def getsons3(self):
        return self.__Sons3

    # Function to calculate the points
    def calculatepoints(self):
        points = 0
        sb = load_workbook(filepath)
        sheet = sb["Sells"]
        row_count = sheet.max_row

        # Code to search the sells the sons in the sheet "Sells"
        for i in range(2, row_count + 1):
            if sheet.cell(row=i, column=1).value != "":
                if sheet.cell(row=i, column=3).value == "Self":  # Self: Sells with own lead
                    data = str(sheet.cell(row=i, column=1).value)
                    if data == self.Name:     # Self sell
                        points = points + 1   # Here you can change the points
                    elif data in self.__Sons1:  # First referal - level 1
                        points = points + 0.5
                    elif data in self.__Sons2:  # Second referal - level 2
                        points = points + 0.25
                    elif data in self.__Sons3:  # Third referal - level 3
                        points = points + 0.125

                elif sheet.cell(row=i, column=3).value == "Lead":  # Lead: Sell with a external lead
                    data = str(sheet.cell(row=i, column=1).value)
                    if data == self.Name:
                        points = points + 0.5
                    elif data in self.__Sons1:
                        points = points + 0.25
                    elif data in self.__Sons2:
                        points = points + 0.125
                    elif data in self.__Sons3:
                        points = points + 0.0625

                # In this case the division of the points is 50-50 but you can change it
                elif sheet.cell(row=i, column=3).value == "Shared":  # Shared: Sell with a shared lead
                    data = str(sheet.cell(row=i, column=1).value)
                    if data == self.Name:
                        points = points + 0.5
                        continue
                    elif data in self.__Sons1:
                        points = points + 0.25
                        continue
                    elif data in self.__Sons2:
                        points = points + 0.125
                        continue
                    elif data in self.__Sons3:
                        points = points + 0.0625
                        continue

                    data2 = str(sheet.cell(row=i, column=2).value)
                    if data2 == self.Name:
                        points = points + 0.5
                    elif data2 in self.__Sons1:
                        points = points + 0.25
                    elif data2 in self.__Sons2:
                        points = points + 0.125
                    elif data2 in self.__Sons3:
                        points = points + 0.0625

        # Code to confirm if the agent is already in the sheet "Points"
        confirm = []
        sheet = sb["Points"]
        row_count = sheet.max_row

        for i in range(2, row_count + 1):
            confirm.append(str(sheet.cell(row=i, column=1).value))

        if self.Name in confirm:
            line = int(confirm.index(self.Name)) + 2
            sheet.cell(row=line, column=2).value = points
            sb.save(filepath)

        else:
            sheet.cell(row=row_count + 1, column=1).value = self.Name
            sheet.cell(row=row_count + 1, column=2).value = points
            sb.save(filepath)


# Code to calculate the point of all agents in the sheet "Agents"
wb = load_workbook(filepath)
sheet = wb["Agents"]
row_count = sheet.max_row
list_agents = []
list_names = []

# Create agent´s objects
for i in range(2, row_count + 1):
    name = sheet.cell(row=i, column=1).value
    if name not in list_names:
        p = Agent(name)
        list_agents.append(p)
        list_names.append(name)

sheet = wb["Referals"]
row_count = sheet.max_row

# Save the first line sons
for i in range(2, row_count + 1):
    name = sheet.cell(row=i, column=1).value
    son = sheet.cell(row=i, column=2).value
    position = list_names.index(name)
    list_agents[position].changesons1(son)

# Save the second line sons
for i in list_agents:
    list_sons = i.getsons1()

    for j in list_sons:
        try:
            o = list_agents[list_names.index(j)]
            o_sons = o.getsons1()

            for k in o_sons:
                i.changesons2(k)

        except ValueError:
            continue

# Save the third line sons
for i in list_agents:
    list_sons = i.getsons1()

    for j in list_sons:
        try:
            l = list_agents[list_names.index(j)]
            list_sons2 = l.getsons1()

        except ValueError:
            continue

        for k in list_sons2:
            try:
                o = list_agents[list_names.index(k)]
                o_sons = o.getsons1()

                for z in o_sons:
                    i.changesons3(z)

            except ValueError:
                continue

# Calculate points for each agent \(._.)/
for i in list_agents:
    print(i.Name)
    print(i.getsons1())
    print(i.getsons2())
    print(i.getsons3())
    i.calculatepoints()
